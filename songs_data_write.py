import os
import eyed3 as e3
from openpyxl import Workbook
from openpyxl import load_workbook

def songs_write(root_path):
    
    wb = load_workbook(filename = root_path + "songs.xlsx")
    ws = wb['Songs']

    n_songs = int(ws["C1"].value)

    for row in ws.iter_rows(min_row=4, max_row=(n_songs + 3), min_col=1, max_col=6):
        song_info = list([cell.value for cell in row])
        song_file = e3.load(root_path + song_info[0])

        song_file.tag.title = song_info[2]
        song_file.tag.artist = song_info[3]
        song_file.tag.album = song_info[4]
        song_file.tag.album_artist = song_info[5]
        song_file.tag.save()

        if song_info[0] != song_info[1] and len(song_info[1]) > 0:
            os.rename(root_path + song_info[0], root_path + song_info[1])
            row[0].value = row[1].value
            wb.save(root_path + "songs.xlsx")
    
    print("Songs successfully modified!")